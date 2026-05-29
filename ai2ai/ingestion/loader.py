"""Source loading, repository mapping, archive handling, and secret hygiene."""

from __future__ import annotations

import base64
import mimetypes
import os
import shutil
import sys
import tarfile
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

from ai2ai.ingestion.chunker import chunk_text_for_ingestion
from ai2ai.ingestion.repo_map import build_repo_map, importance_score
from ai2ai.ingestion.secret_scrubber import RedactionStats, scrub_text
from ai2ai.utils.file_utils import (
    EXCEL_EXT,
    IMG_EXT,
    PDF_EXT,
    TEXT_EXT,
    WORD_EXT,
    ZIP_EXT,
    _excluded,
    decode_text,
    is_probably_binary,
    is_excluded_rel,
    language_for_path,
    normalized_rel,
)


@dataclass
class DocItem:
    name: str
    path: str
    text: str = ""
    b64: str = ""
    mime: str = ""
    kind: str = "text"
    metadata: dict = field(default_factory=dict)

    def is_img(self) -> bool:
        return self.kind == "image"


LAST_INGESTION_METADATA: dict = {}
DEFAULT_MAX_SOURCE_FILES = 1000
DEFAULT_MAX_NESTED_ARCHIVE_DEPTH = 2
DEFAULT_MAX_ARCHIVE_BYTES = 25 * 1024 * 1024


def get_last_ingestion_metadata() -> dict:
    return dict(LAST_INGESTION_METADATA)


def _read_text_path(path: Path) -> str:
    return decode_text(path.read_bytes())


def _read_docx(path: Path) -> str:
    try:
        from docx import Document as DocxReader

        doc = DocxReader(str(path))
        parts = [par.text for par in doc.paragraphs if par.text.strip()]
        for table in doc.tables:
            for row in table.rows:
                rendered = " | ".join(cell.text.strip() for cell in row.cells if cell.text.strip())
                if rendered:
                    parts.append(rendered)
        return "\n\n".join(parts)
    except Exception as exc:
        return f"[DOCX read error: {exc}]"


def _read_pdf(path: Path) -> str:
    try:
        import fitz

        return "\n\n".join(f"[{i + 1}]\n{page.get_text()}" for i, page in enumerate(fitz.open(path)))
    except ImportError:
        pass
    try:
        import pypdf

        reader = pypdf.PdfReader(str(path))
        return "\n\n".join(f"[{i + 1}]\n{page.extract_text()}" for i, page in enumerate(reader.pages))
    except ImportError:
        return "[PDF: pip install pymupdf]"


def _read_excel(path: Path) -> str:
    try:
        import pandas as pd

        ext = path.suffix.lower()
        if ext in (".csv", ".tsv"):
            return pd.read_csv(str(path), encoding_errors="replace").to_string(index=False, max_rows=100)
        dfs = pd.read_excel(str(path), sheet_name=None)
        return "\n\n".join(
            f"--- {sheet} ---\n{df.to_string(index=False, max_rows=50)}"
            for sheet, df in dfs.items()
        )
    except ImportError:
        pass
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            rows = [
                " | ".join(str(cell) if cell else "" for cell in row)
                for row in wb[sheet].iter_rows(max_row=100, values_only=True)
            ]
            parts.append(f"--- {sheet} ---\n" + "\n".join(row for row in rows if row.strip()))
        return "\n\n".join(parts)
    except Exception:
        return "[Excel: pip install openpyxl]"


def _safe_tmp_file(data: bytes, suffix: str) -> Path:
    fd, tmp_name = tempfile.mkstemp(prefix="debate_ingest_", suffix=suffix)
    os.close(fd)
    path = Path(tmp_name)
    path.write_bytes(data)
    return path


def _text_from_bytes(data: bytes, display_path: str, source_path: Path | None = None) -> str:
    ext = Path(display_path).suffix.lower()
    if source_path and ext in WORD_EXT:
        return _read_docx(source_path)
    if source_path and ext in PDF_EXT:
        return _read_pdf(source_path)
    if source_path and ext in EXCEL_EXT:
        return _read_excel(source_path)
    if ext in WORD_EXT | PDF_EXT | EXCEL_EXT:
        tmp = _safe_tmp_file(data, ext)
        try:
            return _text_from_bytes(data, display_path, tmp)
        finally:
            tmp.unlink(missing_ok=True)
    return decode_text(data)


class IngestionSession:
    def __init__(self, max_chars: int, max_source_files: int | None,
                 max_source_chars: int | None, max_nested_archive_depth: int):
        self.max_chars = max_chars
        self.max_source_files = max_source_files or DEFAULT_MAX_SOURCE_FILES
        self.max_source_chars = max_source_chars
        self.max_nested_archive_depth = max_nested_archive_depth
        self.items: list[DocItem] = []
        self.records: list[dict] = []
        self.skipped_files: list[dict] = []
        self.truncated_files: list[dict] = []
        self.nested_archives_processed: list[dict] = []
        self.warnings: list[str] = []
        self.total_chars_before_limits = 0
        self.total_chars_after_limits = 0
        self.redaction_stats = RedactionStats()
        self.source_count = 0

    def skip(self, path: str, reason: str, size: int | None = None):
        self.skipped_files.append({"path": normalized_rel(path), "reason": reason, "size": size})

    def _file_limit_reached(self) -> bool:
        return len(self.items) >= self.max_source_files

    def add_image(self, display_path: str, data: bytes):
        if self._file_limit_reached():
            self.skip(display_path, "max_source_files_limit", len(data))
            return
        mime = mimetypes.guess_type(display_path)[0] or "image/png"
        item = DocItem(
            display_path,
            display_path,
            b64=base64.standard_b64encode(data).decode(),
            mime=mime,
            kind="image",
            metadata={"language": "image", "chars_before": 0, "chars_after": 0},
        )
        self.items.append(item)
        self.records.append({"display_path": display_path, "kind": "image", "language": "image"})

    def add_text(self, display_path: str, data: bytes, source_path: Path | None = None):
        if self._file_limit_reached():
            self.skip(display_path, "max_source_files_limit", len(data))
            return
        if is_probably_binary(data) and Path(display_path).suffix.lower() not in WORD_EXT | PDF_EXT | EXCEL_EXT:
            self.skip(display_path, "binary_file", len(data))
            return
        raw_text = _text_from_bytes(data, display_path, source_path)
        self.total_chars_before_limits += len(raw_text)
        scrubbed, _ = scrub_text(raw_text, self.redaction_stats)
        language = language_for_path(display_path)
        score, reason = importance_score(display_path)
        chunked, chunk_meta = chunk_text_for_ingestion(display_path, scrubbed, self.max_chars)

        if self.max_source_chars is not None and self.total_chars_after_limits + len(chunked) > self.max_source_chars:
            remaining = max(0, self.max_source_chars - self.total_chars_after_limits)
            if remaining <= 0:
                self.skip(display_path, "max_total_text_chars_limit", len(data))
                return
            chunked, chunk_meta = chunk_text_for_ingestion(display_path, scrubbed, remaining)
            chunk_meta["truncated"] = True
            chunk_meta["truncation_reason"] = "max_total_text_chars_limit"

        self.total_chars_after_limits += len(chunked)
        if chunk_meta.get("truncated"):
            self.truncated_files.append(
                {
                    "path": display_path,
                    "reason": chunk_meta.get("truncation_reason", "truncated"),
                    "chars_before": chunk_meta.get("chars_before", len(raw_text)),
                    "chars_after": chunk_meta.get("chars_after", len(chunked)),
                }
            )

        item = DocItem(
            display_path,
            display_path,
            text=chunked,
            kind="text",
            metadata={
                **chunk_meta,
                "language": language,
                "importance_score": score,
                "importance_reason": reason,
            },
        )
        self.items.append(item)
        self.records.append(
            {
                "display_path": display_path,
                "kind": "text",
                "language": language,
                "importance_score": score,
                "importance_reason": reason,
                "text_for_detection": scrubbed[:20000],
            }
        )

    def metadata(self) -> dict:
        repo_map = build_repo_map(
            self.records,
            self.skipped_files,
            self.truncated_files,
            limits={
                "max_chars_per_file": self.max_chars,
                "max_source_files": self.max_source_files,
                "max_source_chars": self.max_source_chars,
                "max_nested_archive_depth": self.max_nested_archive_depth,
            },
        )
        text_count = sum(1 for item in self.items if item.kind == "text")
        image_count = sum(1 for item in self.items if item.kind == "image")
        return {
            "source_count": self.source_count,
            "text_file_count": text_count,
            "image_file_count": image_count,
            "skipped_file_count": len(self.skipped_files),
            "total_chars_before_limits": self.total_chars_before_limits,
            "total_chars_after_limits": self.total_chars_after_limits,
            "repo_map": repo_map,
            "detected_languages": repo_map.get("detected_languages", []),
            "detected_frameworks": repo_map.get("detected_frameworks", []),
            "important_files": repo_map.get("important_source_files", []),
            "skipped_files": self.skipped_files,
            "truncated_files": self.truncated_files,
            "nested_archives_processed": self.nested_archives_processed,
            "redaction_summary": self.redaction_stats.to_dict(),
            "ingestion_warnings": self.warnings,
        }


def _collect_folder(session: IngestionSession, root: Path):
    candidates = []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        if _excluded(path, root):
            session.skip(str(path.relative_to(root)), "excluded_by_default_rules", path.stat().st_size)
            continue
        rel = normalized_rel(path.relative_to(root))
        candidates.append((importance_score(rel)[0], rel, path))
    for _, rel, path in sorted(candidates, key=lambda row: (-row[0], row[1])):
        _process_path(session, path, rel)


def _process_path(session: IngestionSession, path: Path, display_path: str):
    ext = path.suffix.lower()
    try:
        data = path.read_bytes()
    except Exception as exc:
        session.skip(display_path, f"read_error: {exc}")
        return
    if ext in ZIP_EXT:
        _process_archive_bytes(session, data, display_path, 0)
    elif ext in IMG_EXT:
        session.add_image(display_path, data)
    else:
        session.add_text(display_path, data, path)


def _safe_archive_member(name: str) -> bool:
    p = Path(name)
    return not p.is_absolute() and ".." not in p.parts


def _process_archive_bytes(session: IngestionSession, data: bytes, archive_ref: str, depth: int):
    if depth > session.max_nested_archive_depth:
        session.skip(archive_ref, "nested_archive_depth_limit", len(data))
        return
    if len(data) > DEFAULT_MAX_ARCHIVE_BYTES:
        session.skip(archive_ref, "archive_size_limit", len(data))
        return
    session.nested_archives_processed.append({"path": archive_ref, "depth": depth, "size": len(data)})
    suffix = Path(archive_ref).suffix.lower()
    tmp = _safe_tmp_file(data, suffix or ".zip")
    try:
        if suffix in {".tar", ".gz", ".tgz"}:
            with tarfile.open(tmp, "r:*") as archive:
                entries = []
                for member in archive.getmembers():
                    if not member.isfile() or not _safe_archive_member(member.name):
                        continue
                    if is_excluded_rel(member.name):
                        session.skip(f"{archive_ref}::{normalized_rel(member.name)}", "excluded_by_default_rules", member.size)
                        continue
                    entries.append((importance_score(member.name)[0], member.name, member))
                for _, name, member in sorted(entries, key=lambda row: (-row[0], row[1])):
                    extracted = archive.extractfile(member)
                    if not extracted:
                        continue
                    member_data = extracted.read()
                    _process_archive_member(session, archive_ref, name, member_data, depth)
        else:
            with zipfile.ZipFile(tmp) as archive:
                entries = []
                for info in archive.infolist():
                    if info.is_dir() or not _safe_archive_member(info.filename):
                        continue
                    if is_excluded_rel(info.filename):
                        session.skip(f"{archive_ref}::{normalized_rel(info.filename)}", "excluded_by_default_rules", info.file_size)
                        continue
                    entries.append((importance_score(info.filename)[0], info.filename, info))
                for _, name, info in sorted(entries, key=lambda row: (-row[0], row[1])):
                    member_data = archive.read(info)
                    _process_archive_member(session, archive_ref, name, member_data, depth)
    except Exception as exc:
        session.skip(archive_ref, f"archive_read_error: {exc}", len(data))
    finally:
        tmp.unlink(missing_ok=True)


def _process_archive_member(session: IngestionSession, archive_ref: str, member_name: str,
                            data: bytes, depth: int):
    display = f"{archive_ref}::{normalized_rel(member_name)}"
    ext = Path(member_name).suffix.lower()
    if ext in ZIP_EXT:
        if depth >= session.max_nested_archive_depth:
            session.skip(display, "nested_archive_depth_limit", len(data))
            return
        _process_archive_bytes(session, data, display, depth + 1)
    elif ext in IMG_EXT:
        session.add_image(display, data)
    else:
        session.add_text(display, data)


def _print_summary(items: list[DocItem], metadata: dict):
    print(f"\n📂 {len(items) + metadata.get('skipped_file_count', 0)} fájl...")
    for item in items:
        icon = "🖼️ " if item.is_img() else "✅"
        ext = Path(item.name.split("::")[-1]).suffix.upper() or "?"
        size = f"({len(item.text):,} kar)" if not item.is_img() else ""
        print(f"   {icon} {ext:6s} {item.name} {size}")
    imgs = metadata.get("image_file_count", 0)
    print(f"\n   📊 {metadata.get('text_file_count', 0)} szöveges + {imgs} kép")
    if metadata.get("skipped_file_count"):
        print(f"   ⚠️  {metadata['skipped_file_count']} kihagyott fájl")
    redactions = metadata.get("redaction_summary", {}).get("total_redactions", 0)
    if redactions:
        print(f"   🔒 {redactions} titok maszkolva")


def load_sources(sources, max_chars, max_source_files: int | None = None,
                 max_source_chars: int | None = None,
                 max_nested_archive_depth: int = DEFAULT_MAX_NESTED_ARCHIVE_DEPTH,
                 show_summary: bool = False):
    global LAST_INGESTION_METADATA
    session = IngestionSession(max_chars, max_source_files, max_source_chars, max_nested_archive_depth)
    tmp_dirs: list[str] = []
    session.source_count = len(sources or [])
    for src in sources or []:
        path = Path(src)
        if not path.exists():
            print(f"   ⚠️  Nem találom: {src}")
            session.skip(str(src), "source_not_found")
            continue
        if path.is_dir():
            _collect_folder(session, path)
        elif path.suffix.lower() in ZIP_EXT:
            print(f"   🗜️  ZIP: {path.name}")
            _process_path(session, path, path.name)
        else:
            _process_path(session, path, path.name)
    for tmp in tmp_dirs:
        shutil.rmtree(tmp, ignore_errors=True)
    LAST_INGESTION_METADATA = session.metadata()
    if show_summary:
        from ai2ai.ingestion.source_summary import format_ingestion_summary_markdown

        print(format_ingestion_summary_markdown(LAST_INGESTION_METADATA))
    _print_summary(session.items, LAST_INGESTION_METADATA)
    if not session.items:
        print("❌ Nincs feldolgozható fájl.")
        sys.exit(1)
    return session.items
